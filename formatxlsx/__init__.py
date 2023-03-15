import logging
import os.path
import re
from multiprocessing.pool import Pool
from os import walk
from os.path import exists, isfile, isdir, join

import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter


class FMTXlsx(object):
    """paths 解析规则:\r\n
    1.如是文件,文件需以xlsx结尾;\r\n
    2.如是路径,此路径下的所有xlsx结尾的文件;\r\n
    3.以上会被格式化;\r\n
    wrap_text 设置单元格是否自动换行,默认为需要自动换行\r\n
    align 设置单元格对齐方式,默认为居中\r\n
    color 设置边框颜色,默认为red,仅可选值 red， black，green, blue\r\n
    """

    def __init__(self, paths: str, wrap_text=True, align="center", over_name=True, color="1"):
        colors_dict = {"red": "FF0000", "black": "000000", "green": "00FF00", "blue": "0000FF"}
        self.side = Side(style="thin", color=colors_dict.get(color, colors_dict.get("red")))
        self.border = Border(left=self.side, right=self.side, top=self.side, bottom=self.side)
        self.alignments = Alignment(wrap_text=wrap_text, horizontal=align, vertical=align)
        self.paths = paths
        self.repl = re.compile(r"(\.xlsx$)")
        self.repl_suffix = r"fmt\1"
        self.over_name = over_name
        self.ignore_dirs = ["venv", ".idea"]
        self.suffix = "xlsx"
        self.file_names: list = []
        self.parse_paths2file()

    def parse_paths2file(self):
        if not exists(self.paths):
            logging.fatal("遇事不要慌，请找周景棠！！")
        if isfile(self.paths) and self.paths.endswith(self.suffix):
            self.file_names.append(self.paths)
        elif isdir(self.paths):
            for cur_dir, _, files in walk(self.paths):
                for file in files:
                    final_file = join(cur_dir, file)
                    if final_file.endswith(self.suffix) and isfile(final_file):
                        self.file_names.append(final_file)

    """将格式保存到所有的xlsx文件中"""

    def fmt_save(self):
        p = Pool(4)
        for xlsx in self.file_names:
            p.apply_async(self.fmt, args=(xlsx,))
        p.close()
        p.join()

    def fmt(self, file_name):
        dims = {}
        wb = openpyxl.load_workbook(file_name)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            max_column, max_row = ws.max_column, ws.max_row
            for i in range(1, max_column + 1):
                for j in range(1, max_row + 1):
                    blob = ws.cell(row=j, column=i)
                    blob.border = self.border
                    blob.alignment = self.alignments
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                        # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                        # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                        cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                        dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
            for col, value in dims.items():
                # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
                value = 55 if value > 55 else value
                ws.column_dimensions[get_column_letter(col)].width = value + 2
        if not self.over_name:
            d, f = os.path.split(file_name)
            f = self.repl.sub(self.repl_suffix, f)
            file_name = os.path.join(d, f)
        wb.save(file_name)
